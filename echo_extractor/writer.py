"""Write extracted EchoData rows into the xlsx template atomically."""
from __future__ import annotations

import logging
import os
import tempfile
from pathlib import Path
from typing import Any

import openpyxl

from .schema import ALWAYS_BLANK, COLUMN_MAP, EchoData

log = logging.getLogger(__name__)

# Plausible ranges for --verify mode: (min, max) inclusive.
VERIFY_RANGES: dict[str, tuple[float, float]] = {
    "LVIDd": (30.0, 80.0),
    "LVIDs": (20.0, 70.0),
    "IVSd": (6.0, 25.0),
    "LVPWd": (6.0, 25.0),
    "EF": (10.0, 85.0),
    "FS": (10.0, 50.0),
    "EDV": (30.0, 400.0),
    "ESV": (10.0, 300.0),
    "av vmax": (0.5, 6.0),
    "tr vmax": (1.0, 6.0),
    "TAPSE": (5.0, 35.0),
}


def _load_headers(ws) -> dict[str, int]:
    """Return {column_name: 1-based column index} from the header row."""
    return {cell.value: cell.column for cell in ws[1] if cell.value is not None}


def _find_next_empty_row(ws, num_cols: int) -> int:
    """Return the 1-based index of the first fully-empty row starting from row 2."""
    for row_idx in range(2, ws.max_row + 2):
        row_cells = ws[row_idx]
        if all(c.value is None for c in row_cells[:num_cols]):
            return row_idx
    return ws.max_row + 1


def _save_atomic(wb: openpyxl.Workbook, xlsx_path: Path) -> None:
    parent = xlsx_path.parent
    fd, tmp_path = tempfile.mkstemp(dir=parent, suffix=".xlsx.tmp")
    os.close(fd)
    try:
        wb.save(tmp_path)
        os.rename(tmp_path, xlsx_path)
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


def append_patient_row(data: EchoData, xlsx_path: Path, output_path: Path) -> None:
    """Append one row for this patient to output_path, saved atomically."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    headers = _load_headers(ws)
    num_cols = len(headers)
    target_row = _find_next_empty_row(ws, num_cols)

    # Build column_name → value dict from EchoData
    col_values: dict[str, Any] = {}
    for field_name, col_spec in COLUMN_MAP.items():
        value = getattr(data, field_name, None)
        cols = [col_spec] if isinstance(col_spec, str) else col_spec
        for col_name in cols:
            if col_name in ALWAYS_BLANK:
                continue
            col_values[col_name] = value

    # Write values into the target row
    written = 0
    for col_name, value in col_values.items():
        if value is None:
            continue
        col_idx = headers.get(col_name)
        if col_idx is None:
            log.warning("Column %r not found in xlsx headers — skipping", col_name)
            continue
        ws.cell(row=target_row, column=col_idx, value=value)
        written += 1

    log.info("Writing row %d (%d cells filled) → %s", target_row, written, output_path)
    _save_atomic(wb, output_path)


def verify_rows(xlsx_path: Path, max_rows: int = 5) -> None:
    """Print a summary table of extracted values with range flags."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    print(f"\n{'Column':<20} {'Value':>12}  {'Status'}")
    print("-" * 46)

    for row_idx in range(2, min(ws.max_row + 1, 2 + max_rows)):
        row_vals = {headers[i]: ws.cell(row=row_idx, column=i + 1).value for i in range(len(headers))}
        print(f"\n--- Row {row_idx} ---")
        for col, (lo, hi) in VERIFY_RANGES.items():
            val = row_vals.get(col)
            if val is None:
                status = "MISSING"
            elif isinstance(val, (int, float)) and not (lo <= val <= hi):
                status = f"SUSPICIOUS (expected {lo}–{hi})"
            else:
                status = "OK"
            print(f"  {col:<18} {str(val):>12}  {status}")
